#!/usr/bin/env python3
import sys
import json
import re
import logging
import shutil
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from copy import copy
from openpyxl.cell.cell import MergedCell

# -------------------- 로거 --------------------
def setup_logger(log_path: str = "fill_define.log"):
    logger = logging.getLogger("fill_define")
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()

    fmt = logging.Formatter("[%(asctime)s] %(levelname)s: %(message)s", "%H:%M:%S")

    sh = logging.StreamHandler(sys.stdout)
    sh.setLevel(logging.INFO)
    sh.setFormatter(fmt)
    logger.addHandler(sh)

    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)
    logger.addHandler(fh)
    return logger

# -------------------- 진행 표시 --------------------
class Progress:
    def __init__(self, total: int, bar_width: int = 28):
        self.total = max(1, int(total))
        self.count = 0
        self.bar_width = bar_width

    def step(self, msg: str):
        self.count += 1
        pct = int(self.count * 100 / self.total)
        filled = int(self.bar_width * pct / 100)
        bar = "#" * filled + "-" * (self.bar_width - filled)
        print(f"[{bar}] {pct:3d}% - {msg}")

# -------------------- 파일 백업 --------------------
def backup_excel(excel_path: str) -> Path:
    p = Path(excel_path)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"{p.stem}_backup_{ts}{p.suffix}"
    dst = p.with_name(backup_name)
    shutil.copy2(p, dst)
    return dst

# -------------------- 엑셀 헬퍼 --------------------
def find_define_sheet(wb):
    for name in wb.sheetnames:
        if str(name).strip().lower() == "define":
            return wb[name]
    return wb.create_sheet("define")

def find_label_cell(ws, label: str):
    target = label.strip().casefold()
    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            if isinstance(val, str) and val.strip().casefold() == target:
                return cell
    return None

def ensure_base_label(ws, label: str):
    cell = find_label_cell(ws, label)
    if cell:
        return cell
    # 없으면 아래쪽에 라벨 생성(서식은 템플릿 없으면 기본)
    row = ws.max_row + 2
    ws.cell(row=row, column=1).value = label
    return ws.cell(row=row, column=1)

def right_cell(ws, cell):
    return ws.cell(row=cell.row, column=cell.column + 1)

def find_io_header(ws):
    # 한 줄에 Inputs/Bits, Outputs/Bits, Parameters/Bits가 있는 헤더 탐지
    for row in ws.iter_rows():
        header_map = {}
        for cell in row:
            if isinstance(cell.value, str):
                key = cell.value.strip().casefold()
                if key in ("inputs", "outputs", "parameters"):
                    header_map[key] = cell.column
        if header_map.get("inputs") and header_map.get("outputs") and header_map.get("parameters"):
            in_c = header_map["inputs"]
            out_c = header_map["outputs"]
            par_c = header_map["parameters"]
            return {
                "row": row[0].row,
                "inputs_col": in_c, "inputs_bits_col": in_c + 1,
                "outputs_col": out_c, "outputs_bits_col": out_c + 1,
                "params_col": par_c, "params_bits_col": par_c + 1
            }
    return None

def create_io_header(ws):
    # 템플릿 없으면 간단 헤더 생성(서식 없음)
    start_row = ws.max_row + 2
    ws.cell(row=start_row, column=1).value = "Inputs"
    ws.cell(row=start_row, column=2).value = "Bits"
    ws.cell(row=start_row, column=4).value = "Outputs"
    ws.cell(row=start_row, column=5).value = "Bits"
    ws.cell(row=start_row, column=7).value = "Parameters"
    ws.cell(row=start_row, column=8).value = "Bits"
    return {
        "row": start_row,
        "inputs_col": 1, "inputs_bits_col": 2,
        "outputs_col": 4, "outputs_bits_col": 5,
        "params_col": 7, "params_bits_col": 8
    }

def next_empty_row(ws, col, start_row):
    r = max(1, start_row)
    maxr = ws.max_row + 2
    while r <= maxr:
        if ws.cell(row=r, column=col).value in (None, ""):
            return r
        r += 1
    return maxr

# ---------- 스타일 복사(서식 보존) ----------
def copy_cell_style(dst, src):
    # src가 셀이고 서식이 없다면 스킵
    if not src or not getattr(src, "has_style", False):
        return
    # 병합 셀이면 스킵(혹은 병합 범위 좌상단 셀로 대체)
    if isinstance(src, MergedCell):
        return
    try:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        # number_format은 문자열이라 복사 불필요
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)
    except TypeError:
        # 일부 버전/상황에서 StyleProxy가 섞여 들어오면 여기로 옴
        # 서식 복사는 생략하고 값만 쓰도록 조용히 통과
        pass


def clone_row_style(ws, template_row: int, target_row: int, cols):
    # 템플릿 행의 각 셀 스타일을 안전하게 복사
    for c in cols:
        tcell = ws.cell(row=target_row, column=c)
        scell = ws.cell(row=template_row, column=c)
        copy_cell_style(tcell, scell)

# ---------- 폭 변환([msb:lsb] -> 비트수 정수 텍스트) ----------
def bits_text(w):
    if w is None:
        return "1"
    s = str(w).strip()
    if s == "":
        return "1"
    m = re.match(r'^\[\s*(\d+)\s*:\s*(\d+)\s*\]$', s)
    if m:
        msb = int(m.group(1))
        lsb = int(m.group(2))
        return str(abs(msb - lsb) + 1)
    m = re.match(r'^\[\s*\d+\s*\]$', s)
    if m:
        return "1"
    if s.isdigit():
        return s
    # 수식/매크로는 원문 유지
    return s

# ---------- 클리어(값만 비우고 서식 유지) ----------
def last_used_row_in_cols(ws, cols, start_row):
    max_row = ws.max_row
    for r in range(max_row, start_row - 1, -1):
        for c in cols:
            if ws.cell(row=r, column=c).value not in (None, ""):
                return r
    return start_row - 1

def clear_base(ws):
    for label in ("Target Name", "Target Path", "Base Clock", "Base Reset"):
        c = find_label_cell(ws, label)
        if c:
            rc = right_cell(ws, c)
            rc.value = None  # 값만 지움(서식 유지)

def clear_io(ws):
    hdr = find_io_header(ws)
    if not hdr:
        return
    start = hdr["row"] + 1
    cols = [
        hdr["inputs_col"], hdr["inputs_bits_col"],
        hdr["outputs_col"], hdr["outputs_bits_col"],
        hdr["params_col"], hdr["params_bits_col"]
    ]
    end = last_used_row_in_cols(ws, cols, start)
    if end < start:
        return
    for r in range(start, end + 1):
        for c in cols:
            ws.cell(row=r, column=c).value = None  # 값만 지움(서식 유지)

# -------------------- 메인 --------------------
def main():
    if len(sys.argv) != 3:
        print("사용법: python fill_define.py <엑셀파일> <JSON파일>")
        sys.exit(1)

    excel_path = sys.argv[1]
    json_path = sys.argv[2]
    logger = setup_logger("fill_define.log")

    # JSON 로드
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    module = data.get("module") or ""
    top_path = data.get("top_path") or ""
    paths = data.get("paths") or []
    clocks = data.get("clocks") or []
    resets = data.get("resets") or []
    params = data.get("parameters") or []
    inputs = data.get("inputs") or []
    outputs = data.get("outputs") or []

    # Target Path: top_path + "." + path(들), top_path 없으면 paths만
    if top_path:
        target_path_str = ",".join(f"{top_path}.{p}" if p else top_path for p in paths) if paths else top_path
    else:
        target_path_str = ",".join(p for p in paths)

    # 워크북 로드
    wb = load_workbook(excel_path)
    ws = find_define_sheet(wb)

    tasks = []

    # 1) 백업
    def run_backup():
        bak = backup_excel(excel_path)
        logger.info("백업 생성: %s", bak)
    tasks.append((run_backup, "백업 생성"))

    # 2) 클리어
    def run_clear():
        clear_base(ws)
        clear_io(ws)
        logger.debug("기존 값 클리어(서식 유지)")
    tasks.append((run_clear, "기존 값 클리어"))

    # 3) IO 헤더 없으면 생성
    if not find_io_header(ws):
        def run_header():
            create_io_header(ws)
            logger.debug("IO 헤더 생성")
        tasks.append((run_header, "IO 헤더 생성"))

    # 4) Base 재기입(항상 덮어씀)
    def run_base_target_name():
        c = ensure_base_label(ws, "Target Name")
        right_cell(ws, c).value = module
        logger.debug("Base 채움: Target Name -> %s", module)
    tasks.append((run_base_target_name, "Base: Target Name"))

    def run_base_target_path():
        c = ensure_base_label(ws, "Target Path")
        right_cell(ws, c).value = target_path_str
        logger.debug("Base 채움: Target Path -> %s", target_path_str)
    tasks.append((run_base_target_path, "Base: Target Path"))

    def run_base_clock():
        c = ensure_base_label(ws, "Base Clock")
        clk = ",".join(x.get("name", "") for x in clocks if x.get("name"))
        right_cell(ws, c).value = clk
        logger.debug("Base 채움: Base Clock -> %s", clk)
    tasks.append((run_base_clock, "Base: Base Clock"))

    def run_base_reset():
        c = ensure_base_label(ws, "Base Reset")
        rst = ",".join(x.get("name", "") for x in resets if x.get("name"))
        right_cell(ws, c).value = rst
        logger.debug("Base 채움: Base Reset -> %s", rst)
    tasks.append((run_base_reset, "Base: Base Reset"))

    # 5) Inputs
    for it in inputs:
        name = it.get("name")
        if not name:
            continue
        bits = bits_text(it.get("width"))
        def run_in(n=name, b=bits):
            hdr = find_io_header(ws) or create_io_header(ws)
            row = next_empty_row(ws, hdr["inputs_col"], hdr["row"] + 1)
            template_row = row - 1 if row - 1 >= hdr["row"] + 1 else hdr["row"] + 1
            clone_row_style(ws, template_row, row, [hdr["inputs_col"], hdr["inputs_bits_col"]])
            ws.cell(row=row, column=hdr["inputs_col"]).value = n
            ws.cell(row=row, column=hdr["inputs_bits_col"]).value = b
            logger.debug("Input 추가: %s (%s)", n, b)
        tasks.append((run_in, f"Inputs: {name}"))

    # 6) Outputs
    for it in outputs:
        name = it.get("name")
        if not name:
            continue
        bits = bits_text(it.get("width"))
        def run_out(n=name, b=bits):
            hdr = find_io_header(ws) or create_io_header(ws)
            row = next_empty_row(ws, hdr["outputs_col"], hdr["row"] + 1)
            template_row = row - 1 if row - 1 >= hdr["row"] + 1 else hdr["row"] + 1
            clone_row_style(ws, template_row, row, [hdr["outputs_col"], hdr["outputs_bits_col"]])
            ws.cell(row=row, column=hdr["outputs_col"]).value = n
            ws.cell(row=row, column=hdr["outputs_bits_col"]).value = b
            logger.debug("Output 추가: %s (%s)", n, b)
        tasks.append((run_out, f"Outputs: {name}"))

    # 7) Parameters
    for it in params:
        if isinstance(it, dict):
            pname = it.get("name") or ""
            pbits = bits_text(it.get("width") or it.get("bits") or "")
        else:
            pname = str(it)
            pbits = ""
        if not pname:
            continue
        def run_pa(n=pname, b=pbits):
            hdr = find_io_header(ws) or create_io_header(ws)
            row = next_empty_row(ws, hdr["params_col"], hdr["row"] + 1)
            template_row = row - 1 if row - 1 >= hdr["row"] + 1 else hdr["row"] + 1
            clone_row_style(ws, template_row, row, [hdr["params_col"], hdr["params_bits_col"]])
            ws.cell(row=row, column=hdr["params_col"]).value = n
            ws.cell(row=row, column=hdr["params_bits_col"]).value = b
            logger.debug("Parameter 추가: %s (%s)", n, b)
        tasks.append((run_pa, f"Parameters: {pname}"))

    # 8) 저장
    def run_save():
        wb.save(excel_path)
        logger.debug("저장 완료: %s", excel_path)
    tasks.append((run_save, "엑셀 저장"))

    # 실행 + 진행률
    prog = Progress(total=len(tasks))
    logger.info("작업 시작 (총 %d 단계)", len(tasks))
    for runner, desc in tasks:
        runner()
        prog.step(desc)
    logger.info("모든 작업 완료")

if __name__ == "__main__":
    main()